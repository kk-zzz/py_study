def handle_data(data, all = []):
  for row in data:
    tmp = []
    for cell in row:
      cell.value and tmp.append(cell.value)
    if tmp:
      all.append(tmp)
  return all
# 引入包
from openpyxl import load_workbook, Workbook
# 打开excel 合并几个打开几个
wb1 = load_workbook('./data/videos.xlsx')
wb2 = load_workbook('./data/movies.xlsx')
# 激活工作簿
sh1 = wb1.active
sh2 = wb2.active
# 定义一个变量用于存储数据
all = []
# 读取数据，存储数据
# for row in sh1.rows:
#   tmp = []
#   for cell in row:
#     cell.value and tmp.append(cell.value)
#   if tmp:
#     all.append(tmp)
# for row in sh2.rows:
#   tmp = []
#   for cell in row:
#     cell.value and tmp.append(cell.value)
#   if tmp:
#     all.append(tmp)

all = handle_data(sh1.rows)
all = handle_data(sh2.rows)

# print(f'all: {all}')
# 保存数据
new_wb = Workbook()
new_sh = new_wb.active
for row in all: # 遍历数据的每一条
  new_sh.append(row)
new_wb.save('./__test__/merge_videos_movies.xlsx')