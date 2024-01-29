'''
1. 将z1，z2，z3 合并文件，并组成一个新文件，并将其输出
'''
def merge_data_func(sh, all_data = []):
  # print(f'max_column: {sh.max_column}')
  # print(f'max_row: {sh.max_row}')
  for row in sh:
    tmp = []
    for cell in row:
      tmp.append(cell.value)
    tmp and all_data.append(tmp)
  return all_data


from openpyxl import load_workbook, Workbook

wb1 = load_workbook('../data/z1_16.xlsx')
wb2 = load_workbook('../data/z2_16.xlsx')
wb3 = load_workbook('../data/z3_16.xlsx')

sh1 = wb1.active
sh2 = wb2.active
sh3 = wb3.active

all_data = merge_data_func(sh1, [])
all_data = merge_data_func(sh2, all_data)
all_data = merge_data_func(sh3, all_data)

print(f'all_data: {all_data}')
print()

# 遍历all_data，新增一行用于存储 最后一列的数据加总
nwb = Workbook()
nsh = nwb.active
sum = 0
max_column = 0
for row in all_data:
  for index, cell in enumerate(row):
    # index == len(row) -1 and print(f'cell: {cell}')
    if index == len(row) -1:
      max_column = index
      sum += cell
  nsh.append(row)
# max_row = len(all_data) - 1
print(f'sum: {sum}, {max_column}')
nrow = []
for i in range(max_column + 1):
  if i == max_column:
    nrow.append(sum)
  else:
    nrow.append('')
# print(f'nrow: {nrow}')
nsh.append(nrow)
nwb.save('../__test__/merge_sales.xlsx')
