def handle_data(data, all = []):
  for row in data:
    tmp = []
    for cell in row:
      # cell.value and tmp.append(cell.value)
      tmp.append(cell.value)
    if tmp:
      all.append(tmp)
  return all

from openpyxl import load_workbook, Workbook
wb1 = load_workbook('../data/sample2.xlsx')
wb2 = load_workbook('../data/sample3.xlsx')

wb1_sheets = wb1.sheetnames
wb2_sheets = wb2.sheetnames

all = []
for sheet in wb1_sheets:
  all = handle_data(wb1[sheet].rows)
for sheet in wb2_sheets:
  all = handle_data(wb2[sheet].rows)

# print(f'all: {all}')
new_wb = Workbook()
new_sh = new_wb.active
for row in all: # 遍历数据的每一条
  new_sh.append(row)
new_wb.save('../__test__/merge_sample_2_3.xlsx')