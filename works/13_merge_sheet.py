'''
1. 将movie中多个工作表的数据，合并到一个工作表中起名为 all_movie
'''
from openpyxl import load_workbook, Workbook
wb = load_workbook('../data/movie2.xlsx')
sh1 = wb['Sheet']
sh2 = wb['Animation']

all = []
for row in sh1.rows:
  tmp_list = []
  for cell in row:
    print(f'sh1-cell: {cell.value}')
    cell.value and tmp_list.append(cell.value)
  all.append(tmp_list)
for row in sh2.rows:
  tmp_list = []
  for cell in row:
    print(f'sh2-cell: {cell.value}')
    cell.value and tmp_list.append(cell.value)
  all.append(tmp_list)
print()
print(f'sh2-all--: {all}')

# save all_movie
sh3 = wb.create_sheet('all_movie')
for row in all:
  sh3.append(row)
wb.save('../__test__/test_all_movie.xlsx')



'''
2. 将movie表中的数据合并成一个新的excel文件，起名为all_movie
'''
wb = Workbook()
sh = wb.active
for row in all:
  sh.append(row)
wb.save('../__test__/all_movie.xlsx')
