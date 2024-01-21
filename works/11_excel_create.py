'''
1. 通过 python 创建 excel 数据，并把多个数据加入到 excel中第一个 sheet中，数据举例：
'''
movies = [
    ["The Grand Adventure", 8.5, "2021-07-15"],
    ["Mystery of the Blue", 7.8, "2019-11-22"],
    ["Lost in the Wild", 8.2, "2020-03-05"],
    ["Rise of a Hero", 9.1, "2022-05-30"],
    ["A Night in Paris", 7.6, "2018-12-14"],
    ["The Last Stand", 8.9, "2021-09-10"],
    ["Beyond the Horizon", 7.4, "2019-06-18"],
    ["Escape from Reality", 8.3, "2020-01-20"],
    ["Dreams of Space", 7.9, "2023-04-12"],
    ["Shadows in the Dark", 8.6, "2021-10-31"],
    ["Whispers of the Past", 8.0, "2018-07-07"],
    ["Chase the Light", 9.3, "2022-08-23"],
    ["The Forgotten Valley", 7.7, "2019-04-17"],
    ["Echoes of Time", 8.4, "2020-02-29"],
    ["Island of Secrets", 8.8, "2021-11-19"],
    ["Under the Moonlight", 7.5, "2018-05-21"]
]

def create_sheet():
  from openpyxl import Workbook
  wb = Workbook()
  sh = wb.active

  for movie in movies:
    # print(f'movie: {movie}')
    sh.append(movie)

  wb.save('../data/movies.xlsx')

create_sheet()

'''
2. 根据不同年限将工作表1中的数据分割
'''
def analyze_movies():
  from openpyxl import load_workbook
  import os
  from datetime import datetime
  if os.path.exists('../data/movies.xlsx'):
    # print('esists')
    wb = load_workbook('../data/movies.xlsx')
    sh = wb.active
    movies_by_year = {}
    for row in sh.rows:
      # print(f'row: {row[2].value}')
      year_obj = datetime.strptime(row[2].value, '%Y-%m-%d')
      year = datetime.strftime(year_obj, '%Y')
      # print(f'year: {year}')
      # print(f'year: {type(year)}')
      
      if year not in movies_by_year:
        movies_by_year[year] = []
      movies_by_year[year].append(row)
      # print(f'tmp: {movies_by_year}')
    
    # 按年份从大到小排序
    sorted_movies_by_year = {year: movies for year, movies in sorted(movies_by_year.items(), key=lambda x: x[0], reverse=True)}
    print(f'sorted_movies_from_big_to_small: {sorted_movies_by_year}')

    for year, rows in sorted_movies_by_year.items():
      # print(f'{year}: {rows}')
      if f'Sheet-{year}' not in wb.sheetnames:
        wb.create_sheet(f'Sheet-{year}')

      year_sheet = wb[f'Sheet-{year}']

      for row in rows:
        values = [cell.value for cell in row]
        year_sheet.append(values)
    
    wb.save('../__test__/test_movies_1.xlsx')

  else:
    create_sheet()

analyze_movies()