'''
1. 创建名为 video 的excel文件，并将自己喜欢的5个番剧保存在文件中，
其中包含番名、主角名称、评分等级、出品国家
'''

def create_video():
  from openpyxl import Workbook
  wb = Workbook()
  sh = wb.active
  rows = [
    ['长歌行', '迪丽热巴，吴磊，刘宇宁，赵露思', 9.1, '中国'],
    ['爱的迫降', '玄彬，孙艺珍', 9.0, '韩国'],
    ['琅琊榜', '胡歌，刘涛，王凯', 9.5, '中国'],
    ['我的室友是九尾狐', '张基龙，李惠利', 8.8, '韩国'],
    ['玉骨遥', '肖战，任敏', 8.8, '中国']
  ]
  for row in rows:
    sh.append(row)
  wb.save('text_video.xlsx')

create_video()


'''
2. 统计video的表格中评分超过 9分的番剧有多少，出品国家为国产的有多少，
并记录到第二个 sheet 表中
'''
def analyze_video():
  from openpyxl import load_workbook
  import os
  rate_count = 0
  china_count = 0
  if os.path.exists('text_video.xlsx'):
    wb = load_workbook('text_video.xlsx')
    sh = wb.active
    for col in sh.columns:
      print(col)
      for item in col:
        if item.value:
          # print(f"Item value: {item.value}, Type: {type(item.value)}")
          rate_count = rate_count + 1 if type(item.value) == int or type(item.value) == float and item.value >= 9.0 else rate_count
          china_count = china_count + 1 if item.value == '中国' else china_count
    print(f'评分超过9分的有：{rate_count}个, 出品国家是国产的有：{china_count}个')
  else:
    create_video()

analyze_video()