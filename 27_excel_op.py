# pip3 install openpyxl

def open():
  from openpyxl import load_workbook
  wb = load_workbook('xls-simple.xlsx')

  sh1 = wb.active
  # 工作表名不可以是 中文
  sh2 = wb['Sheet1']
  sh3 = wb.get_sheet_by_name('Sheet1') # get_sheet_by_name 已废弃，推荐 sh2

  print(sh1 == sh2 == sh3)

def show_sheets():
  from openpyxl import load_workbook
  wb = load_workbook('xls-simple.xlsx')

  print(f'all sheets: {wb.sheetnames}') # all sheets: ['Sheet1']
  for i in wb:
    print(f'each sheet is: {i.title}') # each sheet is: Sheet1

def get_one_value():
  from openpyxl import load_workbook
  wb = load_workbook('xls-simple.xlsx')
  sh1 = wb['Sheet1']
  v1 = sh1.cell(1, 1)
  print(f'value: {v1}') # value: <Cell 'Sheet1'.A1>
  print(f'A1 value: {v1.value}') # A1 value: 异人之下

  v2 = sh1['A2']
  print(f'A2 value: {v2.value}') # A2 value: 咒术回弹

  v3s = sh1['A1:B2']
  print(f'v3s: {v3s}') # v3s: ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>))
  for val in v3s:
    print(f'val in v3s: {val}')
    for vv in val:
      print(f'vv in v3s: {vv.value}')

def get_all_value():
  from openpyxl import load_workbook
  wb = load_workbook('xls-simple.xlsx')

  sh = wb['Sheet1']
  for row in sh.rows:
    # print(f'row in sh1: {row}')
    for col in row:
      # print(f'col in row in sh1: {col}')
      if col.value:
        print(f'{col} value in sh1 is: {col.value}')


open()
show_sheets()
get_one_value()
get_all_value()
'''
row in sh1: (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>, <Cell 'Sheet1'.D1>, <Cell 'Sheet1'.E1>, <Cell 'Sheet1'.F1>, <Cell 'Sheet1'.G1>)
col in row in sh1: <Cell 'Sheet1'.A1>
col in row in sh1: <Cell 'Sheet1'.B1>
col in row in sh1: <Cell 'Sheet1'.C1>
col in row in sh1: <Cell 'Sheet1'.D1>
col in row in sh1: <Cell 'Sheet1'.E1>
col in row in sh1: <Cell 'Sheet1'.F1>
col in row in sh1: <Cell 'Sheet1'.G1>
row in sh1: (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.D2>, <Cell 'Sheet1'.E2>, <Cell 'Sheet1'.F2>, <Cell 'Sheet1'.G2>)
row in sh1: (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.D3>, <Cell 'Sheet1'.E3>, <Cell 'Sheet1'.F3>, <Cell 'Sheet1'.G3>)
row in sh1: (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>, <Cell 'Sheet1'.D4>, <Cell 'Sheet1'.E4>, <Cell 'Sheet1'.F4>, <Cell 'Sheet1'.G4>)
...
'''

'''
<Cell 'Sheet1'.A1> value in sh1 is: 异人之下
<Cell 'Sheet1'.B1> value in sh1 is: 9.8
<Cell 'Sheet1'.C1> value in sh1 is: 宝儿姐、张楚岚
<Cell 'Sheet1'.A2> value in sh1 is: 咒术回弹
<Cell 'Sheet1'.B2> value in sh1 is: 9.2
<Cell 'Sheet1'.C2> value in sh1 is: 五条悟
<Cell 'Sheet1'.A3> value in sh1 is: 黑色四叶草
<Cell 'Sheet1'.B3> value in sh1 is: 9.5
<Cell 'Sheet1'.C3> value in sh1 is: 魔法帝
<Cell 'Sheet1'.F3> value in sh1 is:
'''