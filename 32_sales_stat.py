# 统计excel数据
def column_letter_to_number(letter):
    """将单个列字母转换为列号（从 1 开始）"""
    return ord(letter.upper()) - ord('A') + 1

# 测试函数
column_f = column_letter_to_number('F')
column_g = column_letter_to_number('G')
print('='*10, column_f)
print('='*10, column_g)
# 引包
from openpyxl import load_workbook
# 打开文件
wb = load_workbook('./data/SaleData.xlsx')
# 激活sheet
# sheet = wb['Sales Data']
sheet = wb['SaleData']
# 定义变量保存数据
all_sheets = []
min_row = 2
min_col = 1
# for row in sheet.title: # 此处拿到的是工作表的name
#   print(f'title=={row}')
# 读取sheet
# for rows in sheet.iter_rows(min_row = min_row, min_col=min_col):
row_title = []
for index, rows in enumerate(sheet):
  if index == 0: # excel 第一行标题
    for cell in rows:
      row_title.append(cell.value)
  else:
    tmp = []
    for cell in rows:
      cell.value and tmp.append(cell.value)
    tmp and all_sheets.append(tmp)

print(f'row_title: {row_title}')
# print(f'all_sheets\n: {all_sheets}, {len(all_sheets)}')
# print()

# sum = [0 for _ in range(3)] # 给 list设置默认长度3 且默认值 = 0
sum = 0
new_row = []
# 汇总data
for index, item in enumerate(all_sheets):
  # 检查是否是公式
  last_it = item[len(item) - 1]
  # print(f'last_it: {last_it}, {isinstance(last_it, str)}')
  if isinstance(last_it, str) and last_it.startswith('='):
    try:
      # 提取公式计算的结果
      # 假设公式是乘法（eg: '=F2*G2'）
      formula_parts = last_it[1:].split('*')
      # print(f'formula_parts: {formula_parts}')
      value1 = item[column_letter_to_number(formula_parts[0][0]) - 1]
      value2 = item[column_letter_to_number(formula_parts[1][0]) - 1]
      # print(f'formula_parts_value: {value1}, {value2}')
      # 计算加总
      sum += value1 * value2
    except Exception as e:
      # 处理可能的错误（例如格式不正确）
      print(f'Error in processing formula {last_it}: {e}')
  else:
    sum += last_it
  if index + min_row - 1 == len(all_sheets):
    new_row = ['' for _ in range(len(item))]
    new_row[len(item) - 1] = sum

# save data
all_sheets.insert(0, row_title) # 将row-title 追加到第一条
all_sheets.append(new_row)
# print()
# print(f'new_-data: {all_sheets}')
# export excel
new_sh = wb.create_sheet('Sum Sale Data')
for row in all_sheets:
  new_sh.append(row)

# save new sheet
# wb.save('./__test__/sales_sum.xlsx')
wb.save('./data/SaleData.xlsx')