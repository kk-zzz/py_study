'''
1. open excel
2. active多个sheet
3. 读取数据
  - 合并几个就读几个
  - 整合所有的数据
4. save data
'''
from openpyxl import load_workbook, Workbook
wb = load_workbook('./__test__/test_video_1.xlsx')
# sh1 = wb['Sheet1']
sh1 = wb['少于300W']
sh2 = wb['多于300W']

all = [] # 存储所有的数据

# read sh1
for row in sh1.rows:
  tmp_list = [] # 用于存取一行的数据
  for cell in row:
    # print(f'cell--{cell.value}')
    cell.value and tmp_list.append(cell.value)
  all.append(tmp_list)
print(f'all-sh1-- {all}')

# read sh2
for row in sh2.rows:
  tmp_list = [] # 用于存取一行的数据
  for cell in row:
    # print(f'cell--{cell.value}')
    cell.value and tmp_list.append(cell.value)
  all.append(tmp_list)
print(f'all-sh2-- {all}')

# save data
wb = Workbook()
sh = wb.active

print('='*10, wb == wb)
for row in all:
  sh.append(row)
wb.save('./__test__/test_video_2.xlsx')


