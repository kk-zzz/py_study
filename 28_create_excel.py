
def new_excel_file():
  from openpyxl import Workbook
  wb = Workbook()
  sh1 = wb.active
  sh2 = wb.create_sheet('番剧信息')
  wb.save('new_test1.xls')

def set_value():
  from openpyxl import Workbook
  wb = Workbook()
  sh1 = wb.active
  sh1['A1'] = 'Apple'
  sh1['A2'] = 18
  sh1['A3'] = 'reading'
  sh1['B1'] = 'Cherry'
  sh1['B2'] = 20
  sh1['B3'] = 'hiking'
  wb.save('new_test1.xls')

def set_many_values():
  rows = [
    ['异人之下', 9.8, '宝儿姐，张楚岚'],
    ['咒术回弹', 9.2, '五条悟'],
    ['黑色四叶草', 9.5, '魔法帝']
  ]

  from openpyxl import Workbook
  wb = Workbook()

  for row in rows:
    sh = wb.active
    sh.append(row)
  wb.save('test3.xlsx')

# new_excel_file()
set_value()
set_many_values()