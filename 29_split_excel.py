from openpyxl import load_workbook

wb = load_workbook('./data/videos.xlsx')
sh = wb.active

data1 = [] # 少于300w
data2 = [] # 多于300w

for row in sh.rows:
  # print(f'rows: {row}')
  if row[3].value:
    num = row[3].value
    # print(f'num: {num}')
    if num >= 300:
      data2.append(row)
    else:
      data1.append(row)


data1_sh = wb.create_sheet('少于300W')
data2_sh = wb.create_sheet('多于300W')

for d in data1:
  tmp_list = []
  for t in d:
    print(f'tmp data1: {t.value}')
    if t.value:
      tmp_list.append(t.value)
  data1_sh.append(tmp_list)

for d in data2:
  tmp_list = []
  for t in d:
    print(f'tmp data2: {t.value}')
    if t.value:
      tmp_list.append(t.value)
  data2_sh.append(tmp_list)

wb.save('./__test__/test_video_1.xlsx')