# 统计excel数据
def column_letter_to_number(letter):
    """将单个列字母转换为列号（从 1 开始）"""
    return ord(letter.upper()) - ord('A') + 1

# 测试函数
column_f = column_letter_to_number('F')
column_g = column_letter_to_number('G')
print('='*10, column_f)
print('='*10, column_g)
# 引包
from openpyxl import load_workbook, Workbook
# 打开文件
wb = load_workbook('./data/SaleData.xlsx')
# 激活sheet
sheets = wb.sheetnames
# 定义变量保存数据
all_sheets = [[] for _ in range(len(sheets))]
print(f'all_sheets_length: {all_sheets}, {len(all_sheets)}')
print()
# 读取sheet
for index, sheet in enumerate(sheets):
  for rows in wb[sheet].iter_rows(min_row=2, min_col=1):
    tmp = []
    for cell in rows:
      cell.value and tmp.append(cell.value)
    tmp and all_sheets[index].append(tmp)

print(f'all_sheets\n: {all_sheets}')
print()

# 汇总data
for index, item in enumerate(all_sheets):
  # print(f'all: {index} -- {item}')
  print()
  for it in item:
    # print(f'it: {it}')
    # 检查是否是公式
    last_it = it[len(it) - 1]
    print(f'last_it: {last_it}, {isinstance(last_it, str)}')
    if isinstance(last_it, str) and last_it.startswith('='):
      try:
        # 提取公式计算的结果
        # 假设公式是乘法（eg: '=F2*G2'）
        formula_parts = last_it[1:].split('*')
        # value1 = 
      except Exception as e:
        # 处理可能的错误（例如格式不正确）
        print(f'Error in processing formula {last_it}: {e}')

# save data
# export excel